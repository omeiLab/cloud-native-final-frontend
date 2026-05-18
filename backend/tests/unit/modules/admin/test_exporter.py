"""CSV / xlsx 匯出 tests"""

from datetime import UTC, datetime
from io import BytesIO

from openpyxl import load_workbook

from app.modules.admin.exporter import export_csv, export_xlsx
from app.shared.admin_ref import RegistrationUserView, RegistrationWithUser


def _item(reg_id: str = "01HRGXXXXXXXXXXXXXXXXXXXXX") -> RegistrationWithUser:
    return RegistrationWithUser(
        id=reg_id,
        user=RegistrationUserView(
            employee_id="E1****",
            name="王*明",
            department="研發部",
            site="HSINCHU",
        ),
        session_title="上午場",
        ticket_type_name="員工票",
        status="WON",
        lottery_rank=42,
        created_at=datetime(2026, 5, 4, 10, 0, tzinfo=UTC),
    )


def test_export_csv_has_bom_and_headers() -> None:
    raw = export_csv([_item()])
    text = raw.decode("utf-8")
    # UTF-8 BOM 讓 Excel 認 charset
    assert text.startswith("﻿")
    # header row
    assert "registration_id" in text
    assert "employee_id" in text
    # data
    assert "王*明" in text
    assert "上午場" in text


def test_export_csv_empty_iterable() -> None:
    raw = export_csv([])
    text = raw.decode("utf-8")
    assert "registration_id" in text # header still present
    # 沒 data row 後就只有 header + EOL
    assert text.count("\n") == 1


def test_export_xlsx_can_be_reopened() -> None:
    raw = export_xlsx([_item(), _item("01HOTHERXXXXXXXXXXXXXXXXX")])
    wb = load_workbook(BytesIO(raw), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    # header + 2 data rows
    assert len(rows) == 3
    assert rows[0][0] == "registration_id"
    assert rows[1][0] == "01HRGXXXXXXXXXXXXXXXXXXXXX"
    assert rows[1][2] == "王*明"
