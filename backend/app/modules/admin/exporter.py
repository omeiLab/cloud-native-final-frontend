"""報表匯出 — CSV(stdlib)+ xlsx(openpyxl)。

設計 06 §12.6 對小規模(< 5000 筆)同步生成,大規模可後續改背景任務 +
物件儲存(MinIO)— 本期先做同步路徑(:size guard 在 service 層)。

:對 user 控的欄位(name / event_title 等)套 OWASP CSV injection
防護(`=/+/-/@/Tab/CR` 開頭加單引號 prefix)。
"""

import csv
import io
from collections.abc import Iterable

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.shared.admin_ref import RegistrationUserView, RegistrationWithUser

# 匯出欄位順序(設計 05 §13.6 + 13.8)
_HEADERS = [
    "registration_id",
    "employee_id",
    "name",
    "department",
    "site",
    "session_title",
    "ticket_type_name",
    "status",
    "lottery_rank",
    "created_at",
]

_CSV_INJECTION_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def _sanitize_cell(value: str) -> str:
    """OWASP CSV Injection 防護:`=/+/-/@/Tab/CR` 開頭加 ' prefix"""
    if value and value[0] in _CSV_INJECTION_PREFIXES:
        return "'" + value
    return value


def sanitize_for_export(items: Iterable[RegistrationWithUser]) -> list[RegistrationWithUser]:
    """對 user 控的字串欄位套 _sanitize_cell;回傳新的 item 串
    (frozen DTO 用 model_copy 重建)"""
    out: list[RegistrationWithUser] = []
    for item in items:
        clean_user = RegistrationUserView(
            employee_id=_sanitize_cell(item.user.employee_id),
            name=_sanitize_cell(item.user.name),
            department=_sanitize_cell(item.user.department) if item.user.department else None,
            site=item.user.site,
        )
        out.append(
            item.model_copy(
                update={
                    "user": clean_user,
                    "session_title": _sanitize_cell(item.session_title),
                    "ticket_type_name": _sanitize_cell(item.ticket_type_name),
                }
            )
        )
    return out


def _row_values(item: RegistrationWithUser) -> list[str]:
    return [
        item.id,
        item.user.employee_id,
        item.user.name,
        item.user.department or "",
        item.user.site,
        item.session_title,
        item.ticket_type_name,
        item.status,
        str(item.lottery_rank) if item.lottery_rank is not None else "",
        item.created_at.isoformat(),
    ]


def export_csv(items: Iterable[RegistrationWithUser]) -> bytes:
    """csv 直接 in-memory 生成 — 加 BOM 讓 Excel 正確認 UTF-8"""
    buf = io.StringIO()
    buf.write("﻿") # UTF-8 BOM
    writer = csv.writer(buf)
    writer.writerow(_HEADERS)
    for item in items:
        writer.writerow(_row_values(item))
    return buf.getvalue().encode("utf-8")


def export_xlsx(items: Iterable[RegistrationWithUser]) -> bytes:
    """xlsx 用 openpyxl in-memory 生成"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Registrations"
    ws.append(_HEADERS)
    for item in items:
        ws.append(_row_values(item))
    # 自動調整欄寬(粗略)
    for i in range(1, len(_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 22
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


__all__ = ["export_csv", "export_xlsx"]
